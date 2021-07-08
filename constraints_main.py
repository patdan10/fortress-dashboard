import constraints_db_pull, constraints_additional_calculations, constraints_data_formatter, constraints_factors, constraints_excel_parser, pandas as pd, os, time
import openpyxl, psycopg2
from datetime import datetime, timedelta
import os


def main():
    today = datetime.today()
    #today -= timedelta(1)
    # Get filepath
    fp = "backup" + today.strftime("%d-%b-%y") + ".xlsx"
    fp2 = "backupMax" + today.strftime("%d-%b-%y") + ".xlsx"
    fp3 = "backupMin" + today.strftime("%d-%b-%y") + ".xlsx"
    
    # If filepath exists, just read the excel. So that you only have to actually pull once per day.
    if os.path.exists(fp):
        df, maxes, mins = constraints_excel_parser.get_excel(fp, fp2, fp3)
        listo = list(df.keys())
        return df, listo, maxes, mins

    # Get data from Database
    print("one")
    output = constraints_db_pull.get_data()
    print("two")
    output = constraints_data_formatter.format(output)
    print("three")
    output, listo, maxes, mins = constraints_db_pull.dict_get_more_data(output)
    listo = list(output.keys())
    
    # Additional calculations
    print("FOUR")
    for l in listo:
        constraints_additional_calculations.calculations(output[l])
    
    # Write to Excel file
    if True:
        print("RIGHT THURR")

        newbook = openpyxl.Workbook()
        newbook.save(fp)



        wb = openpyxl.load_workbook(fp)
        
        # Write them all
        writer = pd.ExcelWriter(fp, engine='openpyxl')
        print("NEXT")
        for sheet, frame in output.items():
            writer.sheets = dict((ws.title[:40], ws) for ws in wb.worksheets)  # need this to prevent overwrite
            frame.to_excel(writer, index=False, sheet_name=sheet)

        writer.save()
        print("SAVED")
        
        # convert data to tables
        wb = openpyxl.load_workbook(fp)
        for ws in wb.worksheets:
            mxrow = ws.max_row
            mxcol = ws.max_column
            tab = openpyxl.worksheet.table.Table(displayName=ws.title, ref="A1:" + ws.cell(mxrow, mxcol).coordinate)
            ws.add_table(tab)
        wb.save(fp)
        print("SWAGGER")





        print("RIGHT THURR")

        newbook = openpyxl.Workbook()
        newbook.save(fp2)

        wb = openpyxl.load_workbook(fp2)

        # Write them all
        writer = pd.ExcelWriter(fp2, engine='openpyxl')
        print("NEXT")
        for sheet, frame in maxes.items():
            writer.sheets = dict((ws.title[:40], ws) for ws in wb.worksheets)  # need this to prevent overwrite
            frame.to_excel(writer, index=False, sheet_name=sheet)

        writer.save()
        print("SAVED")

        # convert data to tables
        wb = openpyxl.load_workbook(fp2)
        for ws in wb.worksheets:
            mxrow = ws.max_row
            mxcol = ws.max_column
            tab = openpyxl.worksheet.table.Table(displayName=ws.title, ref="A1:" + ws.cell(mxrow, mxcol).coordinate)
            ws.add_table(tab)
        wb.save(fp2)
        print("SWAGGER")




        print("RIGHT THURR")

        newbook = openpyxl.Workbook()
        newbook.save(fp3)

        wb = openpyxl.load_workbook(fp3)

        # Write them all
        writer = pd.ExcelWriter(fp3, engine='openpyxl')
        print("NEXT")
        for sheet, frame in mins.items():
            writer.sheets = dict((ws.title[:40], ws) for ws in wb.worksheets)  # need this to prevent overwrite
            frame.to_excel(writer, index=False, sheet_name=sheet)

        writer.save()
        print("SAVED")

        # convert data to tables
        wb = openpyxl.load_workbook(fp3)
        for ws in wb.worksheets:
            mxrow = ws.max_row
            mxcol = ws.max_column
            tab = openpyxl.worksheet.table.Table(displayName=ws.title, ref="A1:" + ws.cell(mxrow, mxcol).coordinate)
            ws.add_table(tab)
        wb.save(fp3)
        print("SWAGGER")

    
    # Return gotten or read data
    return output, listo, maxes, mins
