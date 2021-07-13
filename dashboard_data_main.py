import dashboard_db_pull, pandas as pd, os, time
import openpyxl, psycopg2
from datetime import datetime, timedelta
import os


def main():
    today = datetime.today()
    #today -= timedelta(1)
    # Get filepath
    testfp = "backup_3_"+today.strftime("%d-%b-%y") + ".pkl"
    items = []
    
    # If filepath exists, just read the excel. So that you only have to actually pull once per day.
    if os.path.exists(testfp):
        for i in range(4):
            fp = "backup_" + str(i) + "_"+today.strftime("%d-%b-%y") + ".pkl"
            items.append(pd.read_pickle(fp))
        return items


    # Get data from Database
    print("one")
    output = dashboard_db_pull.get_con_data()
    print("two")
    output = dashboard_data_formatter.format(output)
    print("three")
    output = dashboard_db_pull.get_more_nodes_data(output)
    print("FOUR")
    nodes_info = dashboard_db_pull.get_nodes_all()
    print("FIVE")
    weather = dashboard_db_pull.get_weather()
    print("SIX")
    loads = dashboard_db_pull.get_loads()
    print("SEVEN")


    items = [output, nodes_info, weather, loads]

    # Write to Excel file
    if True:
        for i in range(len(items)):
            fp = "backup_" + str(i) + "_"+today.strftime("%d-%b-%y") + ".pkl"
            items[i].to_pickle(fp)

    # Return gotten or read data
    return items











#OUTDATED
def excel_saver(fp, toWrite):
    print("RIGHT THURR")

    newbook = openpyxl.Workbook()
    newbook.save(fp)

    wb = openpyxl.load_workbook(fp)

    # Write them all
    writer = pd.ExcelWriter(fp, engine='openpyxl')
    print("NEXT")
    for sheet, frame in toWrite.items():
        writer.sheets = dict((ws.title[:40], ws) for ws in wb.worksheets)  # need this to prevent overwrite
        frame.to_excel(writer, index=False, sheet_name=sheet)

    writer.save()
    print("SAVED")

    # convert data to tables
    wb = openpyxl.load_workbook(fp)
    for ws in wb.worksheets:
        mxrow = ws.max_row
        mxcol = ws.max_column
        tab = openpyxl.worksheet.table.Table(displayName=ws.title, ref="A1:" + ws.cell(mxrow, mxcol).coordinate)
        ws.add_table(tab)
    wb.save(fp)
    print("SWAGGER")
